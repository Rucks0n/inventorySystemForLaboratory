"""
app/services/importar_excel.py
-------------------------------
Importa los datos históricos de los archivos Excel originales a SQLite.

Archivos origen (colocar en data/ antes de ejecutar):
    inventario.xlsx  →  ítems por depósito + lista de docentes
    gestion.xlsx     →  movimientos históricos

Orden de ejecución (importante respetar dependencias):
    1. importar_docentes()    — docentes primero, los movimientos los referencian
    2. importar_depositos()   — ítems por depósito
    3. importar_movimientos() — historial (requiere items y docentes ya cargados)

Uso desde consola (solo se hace una vez en la migración inicial):
    python app/services/importar_excel.py

También puede llamarse desde la UI: Configuración → Importar datos históricos.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__)
))))

import openpyxl
from db.database      import obtener_conexion
from app.utils.logger import log
from config.settings  import EXCEL_INVENTARIO, EXCEL_GESTION
from config.constants import (
    HOJAS_DEPOSITOS, HOJA_DOCENTES,
    HOJA_MOVIMIENTOS,
)


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _str(valor) -> str | None:
    """Convierte un valor de celda a string limpio, o None si está vacío."""
    if valor is None:
        return None
    s = str(valor).strip()
    return s if s else None


def _int(valor, defecto: int = 0) -> int:
    """Convierte un valor de celda a int de forma segura."""
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return defecto


def _verificar_archivo(ruta: str) -> bool:
    if not os.path.exists(ruta):
        log.error(f"Archivo no encontrado: {ruta}")
        print(f"  ✗  No se encontró: {ruta}")
        return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
#  DOCENTES  (hoja 'Docente' de inventario.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def importar_docentes() -> int:
    """
    Lee la hoja 'Docente' de inventario.xlsx.
    Columnas esperadas: Docente | Correo | Motivo

    Retorna el número de registros importados.
    """
    if not _verificar_archivo(EXCEL_INVENTARIO):
        return 0

    conn   = obtener_conexion()
    cursor = conn.cursor()
    total  = 0

    wb = openpyxl.load_workbook(EXCEL_INVENTARIO, read_only=True, data_only=True)

    if HOJA_DOCENTES not in wb.sheetnames:
        log.warning(f"Hoja '{HOJA_DOCENTES}' no encontrada en inventario.xlsx")
        wb.close()
        conn.close()
        return 0

    ws = wb[HOJA_DOCENTES]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue                        # saltar encabezado
        nombre, correo, motivo = row[0], row[1], row[2] if len(row) > 2 else None

        nombre = _str(nombre)
        correo = _str(correo)
        if not nombre or not correo:
            continue

        cursor.execute("""
            INSERT OR IGNORE INTO docentes (nombre, correo, motivo)
            VALUES (?, ?, ?)
        """, (nombre, correo, _str(motivo)))
        total += 1

    conn.commit()
    conn.close()
    wb.close()

    log.info(f"Docentes importados: {total}")
    print(f"  ✔  {total} docentes importados.")
    return total


# ─────────────────────────────────────────────────────────────────────────────
#  DEPÓSITOS / ÍTEMS  (hojas Depósito1, 2, 3 de inventario.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def importar_depositos() -> int:
    """
    Lee las hojas Depósito1, Depósito2, Depósito3 de inventario.xlsx.
    Columnas esperadas: ID Item | Item | Referencia | Cantidad | Ubicación | Observaciones

    Retorna el total de ítems importados.
    """
    if not _verificar_archivo(EXCEL_INVENTARIO):
        return 0

    conn   = obtener_conexion()
    cursor = conn.cursor()
    total  = 0

    wb = openpyxl.load_workbook(EXCEL_INVENTARIO, read_only=True, data_only=True)

    for nombre_hoja in HOJAS_DEPOSITOS:
        if nombre_hoja not in wb.sheetnames:
            log.warning(f"Hoja '{nombre_hoja}' no encontrada — se omite.")
            continue

        # Obtener el id del depósito ya insertado en la migración
        cursor.execute("SELECT id FROM depositos WHERE nombre = ?", (nombre_hoja,))
        fila_dep = cursor.fetchone()
        if not fila_dep:
            log.warning(f"Depósito '{nombre_hoja}' no registrado en BD — se omite.")
            continue
        deposito_id = fila_dep["id"]

        ws = wb[nombre_hoja]
        subtotal = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue                    # saltar encabezado
            # Columnas: ID | Nombre | Referencia | Cantidad | Ubicación | Observaciones
            _, nombre, referencia, cantidad, ubicacion, observaciones = (
                list(row) + [None] * 6
            )[:6]

            nombre = _str(nombre)
            if not nombre:
                continue

            cant = _int(cantidad, 0)

            cursor.execute("""
                INSERT INTO items
                    (deposito_id, nombre, referencia, cantidad_total,
                     cantidad_disp, ubicacion, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                deposito_id,
                nombre,
                _str(referencia),
                cant,
                cant,           # al importar, todo está disponible
                _str(ubicacion),
                _str(observaciones),
            ))
            subtotal += 1

        log.info(f"{nombre_hoja}: {subtotal} ítems importados.")
        print(f"  ✔  {nombre_hoja}: {subtotal} ítems.")
        total += subtotal

    conn.commit()
    conn.close()
    wb.close()

    log.info(f"Total ítems importados: {total}")
    return total


# ─────────────────────────────────────────────────────────────────────────────
#  MOVIMIENTOS  (hoja 'Movimientos' de gestion.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def importar_movimientos() -> int:
    """
    Lee la hoja 'Movimientos' de gestion.xlsx.
    Columnas esperadas:
        Fecha y Hora | Nombre del ítem | Nombre Usuario | Código |
        Tipo Movimiento | Cantidad | Docente responsable | Motivo

    Si un usuario no existe, lo crea automáticamente.
    Retorna el número de movimientos importados.
    """
    if not _verificar_archivo(EXCEL_GESTION):
        return 0

    conn   = obtener_conexion()
    cursor = conn.cursor()
    total  = 0

    wb = openpyxl.load_workbook(EXCEL_GESTION, read_only=True, data_only=True)

    if HOJA_MOVIMIENTOS not in wb.sheetnames:
        log.warning(f"Hoja '{HOJA_MOVIMIENTOS}' no encontrada en gestion.xlsx")
        wb.close()
        conn.close()
        return 0

    ws = wb[HOJA_MOVIMIENTOS]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cols = (list(row) + [None] * 8)[:8]
        fecha, nombre_item, nombre_usuario, codigo, tipo, cantidad, docente, motivo = cols

        nombre_item = _str(nombre_item)
        if not nombre_item:
            continue

        # ── Buscar ítem por nombre (case-insensitive) ──────────────────────
        cursor.execute(
            "SELECT id FROM items WHERE LOWER(nombre) = LOWER(?)",
            (nombre_item,)
        )
        item_row = cursor.fetchone()
        item_id  = item_row["id"] if item_row else None

        # ── Buscar o crear usuario ─────────────────────────────────────────
        usuario_id = None
        codigo_str = _str(codigo)
        if codigo_str:
            cursor.execute("SELECT id FROM usuarios WHERE codigo = ?", (codigo_str,))
            u = cursor.fetchone()
            if u:
                usuario_id = u["id"]
            else:
                nombre_u = _str(nombre_usuario) or "Importado"
                cursor.execute(
                    "INSERT INTO usuarios (nombre, codigo) VALUES (?, ?)",
                    (nombre_u, codigo_str)
                )
                usuario_id = cursor.lastrowid

        # ── Buscar docente por nombre ──────────────────────────────────────
        docente_id  = None
        docente_str = _str(docente)
        if docente_str:
            cursor.execute(
                "SELECT id FROM docentes WHERE LOWER(nombre) = LOWER(?)",
                (docente_str,)
            )
            d = cursor.fetchone()
            if d:
                docente_id = d["id"]

        cursor.execute("""
            INSERT INTO movimientos
                (item_id, usuario_id, docente_id, tipo, cantidad, motivo, fecha)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            item_id,
            usuario_id,
            docente_id,
            _str(tipo) or "Desconocido",
            _int(cantidad, 0),
            _str(motivo),
            _str(fecha),
        ))
        total += 1

    conn.commit()
    conn.close()
    wb.close()

    log.info(f"Movimientos importados: {total}")
    print(f"  ✔  {total} movimientos importados.")
    return total


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTACIÓN COMPLETA
# ─────────────────────────────────────────────────────────────────────────────

def importar_todo():
    """
    Ejecuta la importación completa en el orden correcto.
    Llama solo una vez durante la migración inicial.
    """
    print("\n─── Importando docentes ────────────────────────────")
    importar_docentes()

    print("\n─── Importando ítems de depósitos ──────────────────")
    importar_depositos()

    print("\n─── Importando movimientos históricos ──────────────")
    importar_movimientos()

    print("\n✔  Importación completa finalizada.\n")


if __name__ == "__main__":
    importar_todo()
