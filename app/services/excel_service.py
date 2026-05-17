"""
app/services/excel_service.py
------------------------------
Exportación de datos del sistema a archivos Excel (.xlsx).
Separado de importar_excel.py porque son operaciones diferentes:
    importar_excel.py  →  carga inicial de datos históricos (se hace una vez)
    excel_service.py   →  exportaciones que el equipo usa continuamente

Funciones disponibles:
    exportar_items(items, ruta_destino)
    exportar_prestamos(prestamos, ruta_destino)
    exportar_movimientos(movimientos, ruta_destino)
    exportar_docentes(docentes, ruta_destino)

Dependencia: pip install openpyxl
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.utils.logger import log


# ── Estilos base ───────────────────────────────────────────────────────────────

def _estilo_encabezado():
    """Retorna el estilo para la fila de encabezados."""
    return {
        "font":      Font(bold=True, color="FFFFFF", size=11),
        "fill":      PatternFill("solid", fgColor="2E5090"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }

def _borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _aplicar_encabezados(ws, columnas: list[str]):
    """Escribe la fila de encabezados con estilo."""
    estilo = _estilo_encabezado()
    for col_idx, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font      = estilo["font"]
        celda.fill      = estilo["fill"]
        celda.alignment = estilo["alignment"]
        celda.border    = _borde_fino()

def _ajustar_columnas(ws):
    """Ajusta el ancho de cada columna al contenido más largo."""
    for col in ws.columns:
        max_len = 0
        for celda in col:
            try:
                largo = len(str(celda.value or ""))
                if largo > max_len:
                    max_len = largo
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

def _guardar(wb, ruta: str) -> bool:
    """Guarda el workbook en la ruta indicada. Crea directorios si no existen."""
    try:
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        wb.save(ruta)
        log.info(f"Excel guardado: {ruta}")
        return True
    except Exception as e:
        log.error(f"Error guardando Excel en '{ruta}': {e}")
        return False


# ── Funciones de exportación ───────────────────────────────────────────────────

def exportar_items(items: list[dict], ruta_destino: str) -> bool:
    """
    Exporta la lista de ítems de inventario a un archivo Excel.

    Columnas: Depósito | ID | Nombre | Referencia | Total | Disponible | Prestado | Veces prestado | Ubicación | Observaciones
    """
    if not items:
        log.warning("exportar_items: lista vacía, no se genera el archivo.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.freeze_panes = "A2"

    columnas = ["Depósito", "ID", "Nombre", "Referencia", "Total",
                "Disponible", "Prestado", "Veces prestado", "Ubicación", "Observaciones"]
    _aplicar_encabezados(ws, columnas)

    for item in items:
        ws.append([
            item.get("deposito", ""),
            item.get("id", ""),
            item.get("nombre", ""),
            item.get("referencia", ""),
            item.get("cantidad_total", 0),
            item.get("cantidad_disp", 0),
            item.get("cantidad_prest", 0),
            item.get("veces_prestado", 0),
            item.get("ubicacion", ""),
            item.get("observaciones", ""),
        ])

    _ajustar_columnas(ws)
    return _guardar(wb, ruta_destino)


def exportar_prestamos(prestamos: list[dict], ruta_destino: str) -> bool:
    """
    Exporta la lista de préstamos activos (o el historial) a Excel.

    Columnas: ID | Ítem | Referencia | Cantidad | Usuario | Código | Teléfono | Docente | Motivo | Fecha | Límite | Estado
    """
    if not prestamos:
        log.warning("exportar_prestamos: lista vacía.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    ws.freeze_panes = "A2"

    columnas = ["ID", "Ítem", "Referencia", "Cantidad", "Usuario", "Código",
                "Teléfono", "Docente", "Motivo", "Fecha préstamo", "Fecha límite", "Estado"]
    _aplicar_encabezados(ws, columnas)

    for p in prestamos:
        ws.append([
            p.get("id", ""),
            p.get("item_nombre", ""),
            p.get("item_ref", ""),
            p.get("cantidad", 0),
            p.get("usuario_nombre", ""),
            p.get("usuario_codigo", ""),
            p.get("usuario_telefono", ""),
            p.get("docente_nombre", ""),
            p.get("motivo", ""),
            p.get("fecha_prestamo", ""),
            p.get("fecha_limite", ""),
            p.get("estado", ""),
        ])

    _ajustar_columnas(ws)
    return _guardar(wb, ruta_destino)


def exportar_movimientos(movimientos: list[dict], ruta_destino: str) -> bool:
    """
    Exporta el historial de movimientos a Excel.

    Columnas: ID | Fecha | Tipo | Ítem | Referencia | Cantidad | Usuario | Código | Docente | Motivo
    """
    if not movimientos:
        log.warning("exportar_movimientos: lista vacía.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.freeze_panes = "A2"

    columnas = ["ID", "Fecha", "Tipo", "Ítem", "Referencia",
                "Cantidad", "Usuario", "Código", "Docente", "Motivo"]
    _aplicar_encabezados(ws, columnas)

    for m in movimientos:
        ws.append([
            m.get("id", ""),
            m.get("fecha", ""),
            m.get("tipo", ""),
            m.get("item_nombre", ""),
            m.get("item_ref", ""),
            m.get("cantidad", 0),
            m.get("usuario_nombre", ""),
            m.get("usuario_codigo", ""),
            m.get("docente_nombre", ""),
            m.get("motivo", ""),
        ])

    _ajustar_columnas(ws)
    return _guardar(wb, ruta_destino)


def exportar_docentes(docentes: list[dict], ruta_destino: str) -> bool:
    """
    Exporta la lista de docentes a Excel.

    Columnas: ID | Nombre | Correo | Motivo habitual | Activo
    """
    if not docentes:
        log.warning("exportar_docentes: lista vacía.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"
    ws.freeze_panes = "A2"

    columnas = ["ID", "Nombre", "Correo", "Motivo habitual", "Activo"]
    _aplicar_encabezados(ws, columnas)

    for d in docentes:
        ws.append([
            d.get("id", ""),
            d.get("nombre", ""),
            d.get("correo", ""),
            d.get("motivo", ""),
            "Sí" if d.get("activo", 1) else "No",
        ])

    _ajustar_columnas(ws)
    return _guardar(wb, ruta_destino)
